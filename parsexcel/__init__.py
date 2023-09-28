#!/usr/bin/env python
import openpyxl


class ParseExcel:
    
    def get_book(self):
        book = openpyxl.load_workbook('./casesexcel/testcases.xlsx')
        
        return book
    
    def get_sheets(self):
        
        #list
        sheets = self.get_book().get_sheet_names()
        
        return sheets
    
    # 获取数据 [{"sheet":"A","data":[]},{"sheet":"B","data":[]}]
    def get_data(self):
        
        # sheet = self.get_book()["登陆"]
        # minrow=sheet.min_row #最小行
        # maxrow=sheet.max_row #最大行
        # mincol=sheet.min_column #最小列
        # maxcol=sheet.max_column # 最大列
        
        # print(minrow,maxrow,mincol,maxcol)
        
        data = []
        col_data = ["no","mod","title","pre _cond","pre_kdt","operate_step","kdt_step","expect_res","priority","responsibler","remarks"]
        
        for index, item in enumerate(self.get_sheets()):
            sheet_min_row = self.get_book()[item].min_row
            sheet_max_row = self.get_book()[item].max_row
            sheet_min_column = self.get_book()[item].min_column
            sheet_max_column = self.get_book()[item].max_column
            data_json = {}
            data_json["sheet"] = item
            data_json["data"] = []
            
            for row in range(sheet_min_row,sheet_max_row+1):
                perrow_list = {}
                perrow_list["data_id"] = row
                for col in range(sheet_min_column, sheet_max_column+1):
                    per = self.get_book()[item].cell(row=row+1, column=col).value if self.get_book()[item].cell(row=row, column=col).value != None else ""
                    #print(col-1)
                    perrow_list[ col_data[col-1] ] = per
                data_json["data"].append( perrow_list )
                    
            data.append(data_json)
        
        return data
    
    
    # 获取数据 [{"sheet":"A",'data_id': 1,'no': '1','mod': '登陆','title': '老板输入正确信息可以成功登录'},{}]
    def get_arrdata(self):
        
        data=[]
        col_data = ["no","mod","title","details","operate_step","kdt_step","veritydef","veritypath","expect_res","priority","responsibler","remarks"]
        
        
        for index, item in enumerate(self.get_sheets()):
            sheet_min_row = self.get_book()[item].min_row
            sheet_max_row = self.get_book()[item].max_row
            sheet_min_column = self.get_book()[item].min_column
            sheet_max_column = self.get_book()[item].max_column
            
                
            for row in range(sheet_min_row,sheet_max_row+1):
                perrow_list = {}
                perrow_list["sheet"] = index+1
                perrow_list["data_id"] = row
                for col in range(sheet_min_column, sheet_max_column+1):
                    per = self.get_book()[item].cell(row=row+1, column=col).value
                    #print(col-1)
                    perrow_list[ col_data[col-1] ] = per if per != None else ""
                if perrow_list["no"] !="" and perrow_list["mod"] !="" and perrow_list["title"] !="": 
                    data.append( perrow_list )
        
        return data